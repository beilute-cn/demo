# ===== 方式1：使用 openpyxl（推荐，支持 .xlsx）=====

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

# 1. 创建新的 Excel 文件
wb = Workbook()
ws = wb.active
ws.title = "销售数据"

# 2. 写入数据
ws['A1'] = "产品名称"
ws['B1'] = "销售额"
ws['C1'] = "数量"

# 写入多行数据
data = [
    ["苹果", 1000, 50],
    ["香蕉", 800, 60],
    ["橙子", 1200, 40],
    ["葡萄", 1500, 30],
]

for row in data:
    ws.append(row)

# 3. 设置样式
# 标题行加粗、背景色
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 4. 设置列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 10

# 5. 添加公式
ws['D1'] = "总计"
ws['D2'] = "=B2*C2"
ws['D3'] = "=B3*C3"
ws['D4'] = "=B4*C4"
ws['D5'] = "=B5*C5"

# 6. 保存文件
wb.save("销售数据.xlsx")
print("Excel 文件已创建：销售数据.xlsx")


# ===== 读取 Excel 文件 =====

# 1. 加载已存在的文件
wb = load_workbook("销售数据.xlsx")
ws = wb.active

# 2. 读取单元格
print(f"A1 的值: {ws['A1'].value}")
print(f"B2 的值: {ws['B2'].value}")

# 3. 遍历所有行
print("\n所有数据：")
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)

# 4. 遍历指定范围
print("\n产品和销售额：")
for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=2, values_only=True):
    print(f"产品: {row[0]}, 销售额: {row[1]}")

# 5. 获取最大行列数
print(f"\n最大行数: {ws.max_row}")
print(f"最大列数: {ws.max_column}")


# ===== 方式2：使用 pandas（数据分析推荐）=====

import pandas as pd

# 1. 创建 DataFrame
df = pd.DataFrame({
    '产品名称': ['苹果', '香蕉', '橙子', '葡萄'],
    '销售额': [1000, 800, 1200, 1500],
    '数量': [50, 60, 40, 30]
})

# 2. 写入 Excel
df.to_excel('销售数据_pandas.xlsx', sheet_name='销售数据', index=False)
print("\nPandas Excel 文件已创建：销售数据_pandas.xlsx")

# 3. 读取 Excel
df_read = pd.read_excel('销售数据_pandas.xlsx')
print("\n读取的数据：")
print(df_read)

# 4. 数据处理
df_read['总计'] = df_read['销售额'] * df_read['数量']
print("\n添加总计列：")
print(df_read)

# 5. 写入多个 sheet
with pd.ExcelWriter('多sheet数据.xlsx') as writer:
    df.to_excel(writer, sheet_name='销售数据', index=False)
    df_read.to_excel(writer, sheet_name='处理后数据', index=False)

print("\n多 sheet 文件已创建：多sheet数据.xlsx")


# ===== 方式3：使用 xlwings（支持宏、图表）=====

# import xlwings as xw
# 
# # 1. 创建新工作簿
# wb = xw.Book()
# ws = wb.sheets['Sheet1']
# 
# # 2. 写入数据
# ws.range('A1').value = [['产品', '销售额'], ['苹果', 1000], ['香蕉', 800]]
# 
# # 3. 读取数据
# data = ws.range('A1:B3').value
# print(data)
# 
# # 4. 保存
# wb.save('xlwings_demo.xlsx')
# wb.close()


# ===== 完整示例：创建带格式的报表 =====

def create_sales_report():
    """创建格式化的销售报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, Reference
    
    wb = Workbook()
    ws = wb.active
    ws.title = "月度销售报表"
    
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = "2024年1月销售报表"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ['产品名称', '单价', '数量', '销售额', '占比']
    ws.append(headers)
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 数据
    data = [
        ['iPhone 15', 5999, 120],
        ['iPad Pro', 6799, 80],
        ['MacBook Air', 7999, 60],
        ['AirPods Pro', 1999, 200],
        ['Apple Watch', 2999, 150],
    ]
    
    # 写入数据并计算
    for row_data in data:
        row = [row_data[0], row_data[1], row_data[2], 
               f"=B{ws.max_row+1}*C{ws.max_row+1}"]
        ws.append(row)
    
    # 添加总计行
    total_row = ws.max_row + 1
    ws[f'A{total_row}'] = "总计"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = f"=SUM(D3:D{total_row-1})"
    
    # 计算占比
    for row in range(3, total_row):
        ws[f'E{row}'] = f"=D{row}/$D${total_row}"
        ws[f'E{row}'].number_format = '0.00%'
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[f'A2:E{total_row}']:
        for cell in row:
            cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    
    # 添加图表
    chart = BarChart()
    chart.title = "产品销售额对比"
    chart.x_axis.title = "产品"
    chart.y_axis.title = "销售额"
    
    data_ref = Reference(ws, min_col=4, min_row=2, max_row=total_row-1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=total_row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    ws.add_chart(chart, "G2")
    
    # 保存
    wb.save("月度销售报表.xlsx")
    print("\n销售报表已创建：月度销售报表.xlsx")

create_sales_report()


# ===== 常用操作汇总 =====

def excel_operations():
    """Excel 常用操作示例"""
    from openpyxl import Workbook, load_workbook
    
    # 1. 创建工作簿
    wb = Workbook()
    
    # 2. 创建新工作表
    ws1 = wb.create_sheet("数据表1")
    ws2 = wb.create_sheet("数据表2", 0)  # 插入到第一个位置
    
    # 3. 删除工作表
    # wb.remove(ws2)
    
    # 4. 复制工作表
    ws3 = wb.copy_worksheet(ws1)
    
    # 5. 获取所有工作表名称
    print("所有工作表:", wb.sheetnames)
    
    # 6. 切换工作表
    ws = wb["数据表1"]
    
    # 7. 插入行/列
    ws.insert_rows(1)  # 在第1行前插入
    ws.insert_cols(1)  # 在第1列前插入
    
    # 8. 删除行/列
    ws.delete_rows(1)
    ws.delete_cols(1)
    
    # 9. 合并单元格
    ws.merge_cells('A1:C1')
    
    # 10. 取消合并
    ws.unmerge_cells('A1:C1')
    
    wb.save("操作示例.xlsx")

excel_operations()
