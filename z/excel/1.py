from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell

# 创建新工作簿
wb = Workbook()
ws = wb.active
ws.title = "示例表"

# 写入数据
ws["A1"] = "姓名"
ws["B1"] = "分数"
ws.append(["张三", 95])
ws.append(["李四", 87])

# 设置样式
ws["A1"].font = Font(bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

# 合并单元格
ws.merge_cells("C2:E5")
# ws["F3"] = "test"


# 保存
wb.save("example.xlsx")


# 读取Excel
wb = load_workbook("example.xlsx")
ws = wb.active

# 读取数据
for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
    print(row)

# 读取特定单元格
print(ws["B2"].value)
