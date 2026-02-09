# 穷举所有合并区域


from openpyxl import load_workbook

# 加载工作簿
wb = load_workbook("example.xlsx")
ws = wb.active


# 方法1: 遍历所有合并单元格范围
def find_merge_top_left(sheet, cell_coordinate):
    """
    找到指定单元格所在合并区域的左上角单元格

    :param sheet: 工作表对象
    :param cell_coordinate: 单元格坐标，如 'C3'
    :return: 左上角单元格坐标，如果不是合并单元格则返回原坐标
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell_coordinate in merged_range:
            # 返回合并区域的左上角单元格坐标
            return merged_range.min_row, merged_range.min_col

    # 如果不在任何合并区域中，返回None或原单元格
    return None


# 使用示例
cell = "C3"
result = find_merge_top_left(ws, cell)

if result:
    row, col = result
    print(f"合并单元格的左上角是: 行{row}, 列{col}")
    # 或者转换为字母坐标
    from openpyxl.utils import get_column_letter

    top_left = f"{get_column_letter(col)}{row}"
    print(f"左上角单元格坐标: {top_left}")

    # 获取左上角单元格的值
    value = ws.cell(row, col).value
    print(f"左上角单元格的值: {value}")
else:
    print(f"{cell} 不是合并单元格")
